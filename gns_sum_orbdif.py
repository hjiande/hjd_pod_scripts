import os
import sys
import xlrd
import xlwt
import glog
import numpy as np
import xml.etree.ElementTree as etree
import pretty_xml
import gns_pod_tool as tool
import openpyxl
from openpyxl import Workbook

def read_one_orbdif_file(file_path):
    glog.fatal("====> read : " + file_path)

    if not os.path.exists(file_path):
        glog.error("The file not exist : " + file_path)
        return

    PRN = ""
    RMS = ""
    LSQ = {}
    ORI = {}
    SAT = []
    try:
        with open(file_path) as file_object:
            # get the sat list
            lines = file_object.readlines()
            for line in lines:
                if "SAT" in line:
                    line = line[line.find("SAT") + 4:]
                    SAT = line.split()
                    for sat in SAT:
                        ORI[sat] = []
                    break
                else:
                    continue

            # get the ORI data
            for line in lines:
                line = line.rstrip()
                if "GPST" in line:
                    continue
                elif "SAT" in line:
                    break
                else:
                    tmp = line.split()
                    ORI[tmp[1]].append(line)
    except:
        glog.error("The file not open : " + file_path)
        sys.exit()
    else:
        glog.info("Os error : " + file_path)

    result = {"ORI": ORI, "SAT": SAT}
    return result


# used for smooth the xml file with the first lsq result
def smooth_with_orbdif(prj_dir, xml_dir, year, doy):
 # jdhuang
    year = '%4d' % int(year)
    doy = '%03d' % int(doy)

    xml = 'gns_pod_%s%s.xml' % (year, doy)
    glog.fatal("====> smooth the xml file : " + xml)
    dif = 'orbdif%s%s' % (year, doy)
    prj_path = os.path.join(prj_dir, dif)
    xml_path = os.path.join(xml_dir, xml)

    orbdif_result = read_one_orbdif_file(prj_path)
    ori_dif = orbdif_result["ORI"]
    # rms ok
    normal_sat = []
    # rms error
    remove_sat = []
    for str_sat in ori_dif:
        A = []
        C = []
        R = []
        sat_ori_dir = ori_dif[str_sat]
        # rms for one sat
        count = 0
        for epo_sat_ori_dir in sat_ori_dir:
            tmp = epo_sat_ori_dir.split()
            a = abs(float(tmp[4]))
            c = abs(float(tmp[5]))
            r = abs(float(tmp[6]))
            A.append(a)
            C.append(c)
            R.append(r)
        mean_a = np.nanmean(A)
        mean_c = np.nanmean(C)
        mean_r = np.nanmean(R)
        # check the mean error
        if mean_a > float(0.40) or mean_c > float(0.30) or mean_r > float(0.30):
            #normal_sat.append(str_sat)
            remove_sat.append(str_sat) 
            print("remove " + str_sat + " " + str(mean_a) + " " + str(mean_c) + " " + str(mean_r))            
        else:
            normal_sat.append(str_sat)

    # 获取tree 对象
    tree = etree.parse(xml_path)
    # 获取根节点
    config = tree.getroot()
    all_sys = ['gps', 'gal', 'bds', 'glo']
    for str_sys in all_sys:
        xml_sys = config.find(str_sys)
        if not xml_sys:
            continue
        xml_sat = xml_sys.find("sat")
        ori_sats = str(xml_sat.text)
        for bad_sat in remove_sat:
            glog.fatal("====> sat is remove for big rms : " + bad_sat)
            ori_sats = ori_sats.replace(bad_sat, "   ")
            xml_sat.text = ori_sats
    config = pretty_xml.prettyXml(config, '\t', '\n')
    tree.write(xml_path, encoding='utf-8', xml_declaration=True)


# print(orbdif_result)


def read_orbdif_files(prj_dir, year, beg_doy, end_doy, post='_fix'):
    tool.check_path(prj_dir)

    doys_rms = {}
    wb_mean = Workbook()
    ws_mean = wb_mean.create_sheet("Mysheet")  # 插入到最后 (默认)
    ws_mean.title = "mean_rms"

    # process not only one day
    for i in range(int(beg_doy), int(end_doy) + 1):
        year = '%4s' % year
        doy = '%03d' % i
        yyyydoy = '%4s%03d' % (year, i)
        dif = "orbdif"+yyyydoy
        doy_path = os.path.join(prj_dir, year, doy, "result")
        tool.check_path(doy_path)

        lsq_dif = os.path.join(doy_path, dif + post)
        doy_flo_result = read_one_orbdif_file(lsq_dif)

        # ======== the sum get form great in orbdif file
        sat_used_today = doy_flo_result["PRN"].split()
        rms_mean_today = doy_flo_result["RMS"].split()
        num = 0
        for prn in sat_used_today:
            if prn == "PRN":
                num = num + 1
                continue
            if prn in doys_rms:
                doys_rms[prn].append(rms_mean_today[num])
            else:
                key = prn
                doys_rms[key] = [key]
                doys_rms[key].append(rms_mean_today[num])
            num = num + 1

        # ======== excel for epoch result without lsq
        wb = Workbook()
        epoch_ori_result = doy_flo_result["ORI"]
        for str_sat in epoch_ori_result:
            ws = wb.create_sheet("Mysheet")  # 插入到最后 (默认)
            ws.title = str_sat
            for line in epoch_ori_result[str_sat]:
                line = line[4:]
                tmp = line.split()
                ws.append(tmp)
        wb.save('float_ori_%s.xlsx' % yyyydoy)

        # ======== excel for epoch result with lsq
        epoch_lsq_result = doy_flo_result["LSQ"]
        wb = Workbook()
        sat_acr = ["A","C","R"]
        for sat in doy_flo_result["SAT"]:
            sat_acr.append(sat+"_A")
            sat_acr.append(sat+"_C")
            sat_acr.append(sat+"_R")
        ws = wb.create_sheet()  # 插入到最后 (默认)
        ws.title = "LSQ_EPOCH"
        ws.append(sat_acr)
        for mjd in epoch_lsq_result:
            line = epoch_lsq_result[mjd]
            tmp = line.split()
            ws.append(tmp)
        wb.save('float_lsq_%s.xlsx' % yyyydoy)


    # =========== mean rms for each sat
    for rms in doys_rms:
        ws_mean.append(doys_rms[rms])
    wb_mean.save("orbdif_" + ('%03d' % beg_doy) + "_" + ('%03d' % end_doy) + ".xlsx")
